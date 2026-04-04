"""
Safety Intelligence Module - Document Generation & AI Analysis
"""
import io
from datetime import datetime
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.colors import HexColor, black
from reportlab.lib.units import inch


def generate_incident_report_excel(posts, project_name, month=None):
    """Generate Excel incident report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Incident Report"
    
    # Styles
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    title_font = Font(bold=True, size=16)
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f'HSE Incident Report - {project_name}'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = f'Period: {month or "All Time"}'
    
    # Headers
    headers = ['ID', 'Date', 'Author', 'Content', 'Status', 'Project', 'Risk Level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, post in enumerate(posts, 6):
        ws.cell(row=row_idx, column=1, value=post.id)
        ws.cell(row=row_idx, column=2, value=post.created_at.strftime('%Y-%m-%d') if post.created_at else 'N/A')
        ws.cell(row=row_idx, column=3, value=post.author.username if post.author else 'Unknown')
        ws.cell(row=row_idx, column=4, value=post.content[:100] if len(post.content) > 100 else post.content)
        ws.cell(row=row_idx, column=5, value=post.status)
        ws.cell(row=row_idx, column=6, value=post.project.name if post.project else 'N/A')
        
        # Simple risk assessment
        risk = 'HIGH' if post.status == 'Pending' else 'LOW'
        ws.cell(row=row_idx, column=7, value=risk)
    
    # Auto-width
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_incident_report_pdf(posts, project_name, month=None):
    """Generate PDF incident report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        textColor='#C00000', fontSize=18, spaceAfter=6
    )
    header_style = ParagraphStyle(
        'CustomHeader', parent=styles['Heading2'],
        textColor='#333333', fontSize=12
    )
    
    # Header
    elements.append(Paragraph(f'HSE Incident Report - {project_name}', title_style))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
    elements.append(Paragraph(f'Period: {month or "All Time"}', styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary Stats
    total = posts.count() if hasattr(posts, 'count') else len(posts)
    open_count = sum(1 for p in posts if p.status == 'Pending')
    closed_count = total - open_count
    
    elements.append(Paragraph('Summary Statistics', header_style))
    stats_data = [
        ['Total Observations', str(total)],
        ['Open/Pending', str(open_count)],
        ['Closed/Resolved', str(closed_count)],
        ['Resolution Rate', f'{(closed_count/total*100) if total else 0:.1f}%']
    ]
    stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), '#f0f0f0'),
        ('TEXTCOLOR', (0, 0), (-1, -1), black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, '#cccccc'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Observations Table
    elements.append(Paragraph('Detailed Observations', header_style))
    table_data = [['ID', 'Date', 'Author', 'Observation', 'Status']]
    for post in posts[:50]:  # Limit to 50
        table_data.append([
            str(post.id),
            (post.created_at.strftime('%m/%d') if post.created_at else 'N/A'),
            (post.author.username[:15] if post.author else 'Unknown'),
            (post.content[:40] + '...' if len(post.content) > 40 else post.content),
            post.status
        ])
    
    obs_table = Table(table_data, colWidths=[0.4*inch, 0.8*inch, 1*inch, 3.3*inch, 0.8*inch])
    obs_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#C00000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#FFF5F5')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, '#dddddd'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFF5F5'), HexColor('#FFFFFF')]),
    ]))
    elements.append(obs_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_risk_trend_ai(posts, project_name, api_key=None):
    """Generate AI-powered risk trend analysis."""
    try:
        import openai
        if not api_key:
            return {'error': 'OpenAI API key not configured'}
        
        client = openai.Client(api_key=api_key)
        
        # Prepare observation data for AI
        observations = []
        for post in posts[:20]:
            observations.append({
                'content': post.content,
                'status': post.status,
                'date': post.created_at.strftime('%Y-%m-%d') if post.created_at else 'Unknown',
                'author': post.author.username if post.author else 'Anonymous'
            })
        
        prompt = f"""Analyze these HSE observations for {project_name} and provide:
1. RISK TREND PREDICTION - What safety risks are trending?
2. AREA-SPECIFIC ANALYSIS - Which areas have highest risk?
3. ACTIONABLE RECOMMENDATIONS - What immediate steps should be taken?

Observations (last 20):
{observations}

Format your response as JSON with keys: risk_trend, high_risk_areas, key_findings, recommendations, predicted_next_month_risk"""

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": "You are an expert HSE safety analyst."},
                     {"role": "user", "content": prompt}],
            temperature=0.3
        )
        
        import json
        try:
            return json.loads(response.choices[0].message.content)
        except:
            return {'analysis': response.choices[0].message.content}
    except ImportError:
        return {'error': 'OpenAI package not installed. Run: pip install openai'}
    except Exception as e:
        return {'error': str(e)}


def generate_hip_ai(task_name, api_key=None):
    """Generate Hazard Identification Plan (HIP) using AI."""
    try:
        import openai
        if not api_key:
            return {'error': 'OpenAI API key not configured'}
        
        client = openai.Client(api_key=api_key)
        
        prompt = f"""Create a comprehensive Hazard Identification Plan (HIP) for this task: {task_name}

Please provide a structured HIP document with:
1. Task Overview
2. Potential Hazards (list minimum 5)
3. Risk Assessment (Severity x Likelihood)
4. Control Measures (Engineering, Administrative, PPE)
5. Emergency Procedures
6. Required Permits/Approvals

Format your response as a structured report suitable for HSE documentation."""

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": "You are an expert HSE professional specializing in hazard identification and risk assessment."},
                     {"role": "user", "content": prompt}],
            temperature=0.3
        )
        
        return {'hip_document': response.choices[0].message.content, 'task': task_name}
    except ImportError:
        return {'error': 'OpenAI package not installed. Run: pip install openai'}
    except Exception as e:
        return {'error': str(e)}